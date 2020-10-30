from jinja2 import Template, Environment, FileSystemLoader
import smtplib
from email.mime.text import MIMEText
import openpyxl
import time

# メールサーバー名
host = "your.host.name"
# 接続先ポート番号
port = 587
# メールサーバーにログインする際のユーザー名とパスワード
user = "your_user"
pwd = "your_pass"
# 送信元のメールアドレス
frm = "your@mail.address"
# メールを1通送信するごとの待ち時間
delay = 1

mail_tmpl_file = "mail.tmpl"
subj_tmpl_file = "subj.tmpl"
data_file = "data.xlsx"

# テンプレートを読み込む
env = Environment(loader=FileSystemLoader("./", encoding = "utf8"))
mail_tmpl = env.get_template(mail_tmpl_file)
subj_tmpl = env.get_template(subj_tmpl_file)

# データファイルの先頭行から項目名を読み込む
wb = openpyxl.load_workbook(data_file)
ws = wb.worksheets[0]
cols = []
col = 1
while ws.cell(1, col).value is not None and ws.cell(1, col).value != "":
    cols.append(ws.cell(1, col).value)
    col += 1
col_count = col

# データファイルの各行を読み込みメールを送信する
row = 2
while ws.cell(row, 1).value is not None and ws.cell(row, 1).value != "":
    # 1行分のデータを読み込む
    params = {}
    for col in range(1, col_count):
        params[cols[col - 1]] = ws.cell(row, col).value

    # テンプレートからメールのタイトルと文章を生成する
    subj = subj_tmpl.render(params)
    body = mail_tmpl.render(params)

    # メールサーバーに接続する
    server = smtplib.SMTP(host, port)
    server.login(user, pwd)

    # メールを送信する
    msg = MIMEText(body)
    msg["Subject"] = subj
    msg["To"] = params["email"]
    msg["From"] = frm
    server.send_message(msg)
    server.quit()
    print("send mail to " + params["email"])

    # 次のメール送信まで時間を空ける
    if delay > 0:
        time.sleep(delay)

    # 次の行に進む
    row += 1
